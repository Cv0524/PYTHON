from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
# wb = Workbook()

#create an active worksheet

wb = load_workbook('C:/PYTHON/py_ex/excel.xlsx')

#print 
ws = wb.active
# print(f'Employee ID {ws["A2"].value} Name: {ws["B2"].value}')

# #Using variables

# employee_id = ws["A3"].value
# name = ws["B3"].value

# print(f'Employee ID {employee_id} Name: {name}')


col = ws["B"]
row = ws['1']
for cell in col:
    print(cell.value)
for row in row:
    print(row.value)
    